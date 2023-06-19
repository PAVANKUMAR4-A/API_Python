import openpyxl, random
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook


class Header_DB_Class():
    random_number = 0
    new_df = pd.DataFrame([])
    combined_df = pd.DataFrame([])


    def generate_random_number(self, length):
        # Define the range of digits (0-9) as a string
        digits = "0123456789"
        prefix = 'SYN_'

        random_number = ''.join(random.choice(digits) for _ in range(length))
        unique_number = prefix+random_number

        return (unique_number)


    def save_header_to_dataset_sheet_db(self,header_dict2,NumOfRecords):

        # writing to 'DataSet' sheet
        df = pd.read_excel(r'sample_synthetic/DB_Tables.xlsx',sheet_name='DataSet')

        column_name = df['DataSet_GUID']
        column_values = column_name.values

        random_number = self.generate_random_number(6)


        if random_number not in column_values:
            header_dict2['DataSet_GUID']=random_number
            #print('DGUID', header_dict2['DataSet_GUID'])

        workbook = openpyxl.load_workbook(r'sample_synthetic/DB_Tables.xlsx')
        sheet = workbook['DataSet']

        row_values = list(header_dict2.values())
        sheet.append(row_values)


        sheet = workbook['DataSet_ENG']
        new_dict = {key:value for key,value in header_dict2.items() if key in ["DataSet_GUID","DataSet_Name"]}
        # new_dict = list(new_dict)*(int(header_dict2["NumOfRecords"]))
        print('New dict data', new_dict)
        row_values = list(new_dict.values()) * NumOfRecords
        print('New row values', row_values)


        dict_list = [new_dict] * NumOfRecords
        df = pd.DataFrame(dict_list)
        df.to_excel(r'sample_synthetic/DB_Tables.xlsx', sheet_name='DataSet_ENG', header=True, index=False)
        print('Latest DF', df)

        sheet.append(row_values)
        workbook.save(r'sample_synthetic/DB_Tables.xlsx')

        # print('hd', header_dict2)





        # dict_list = [new_dict] * NumOfRecords
        # print('Dict list', dict_list)
        #
        # workbook = openpyxl.load_workbook(r'sample_synthetic/DB_Tables.xlsx')
        #
        # # Select the active sheet or create a new sheet
        # sheet = workbook.active if workbook.sheetnames else workbook.create_sheet()
        #
        # # Check if the sheet contains any rows
        # has_data = sheet.max_row > 0
        #
        # # Write the headers if the sheet is empty
        # if not has_data:
        #     headers = list(dict_list[0].keys())
        #     sheet.append(headers)
        #
        # # Write the data row by row
        # for data in dict_list:
        #     row_data = list(data.values())
        #     sheet.append(row_data)







    def save_output_df_dataset_ENG_db(self):

        # Assuming you have a DataFrame called 'df'
        df = pd.read_excel(r'sample_synthetic/concat_dataframe_final.xlsx', sheet_name='Sheet1',index_col=False)
        # print('FIRS DF', df)

        df = df.reset_index(drop=True)

        df = df.drop("Unnamed: 0",axis=1)
        print('dropped df', df)

        # Specify the Excel file path
        excel_file = r'sample_synthetic/DB_Tables.xlsx'

        # Specify the sheet name
        sheet_name = 'DataSet_ENG'

        workbook = openpyxl.load_workbook(excel_file)

        sheet = workbook[sheet_name]

        # Get the last row index in the sheet
        last_row = sheet.max_row
        print('last row data', last_row)

        for col_idx, col_name in enumerate(df.columns, start=3):
            sheet.cell(row=1, column=col_idx, value=col_name)

        for idx, row in df.iterrows():
            for col_idx, value in enumerate(row, start=3):
                sheet.cell(row=last_row + idx, column=col_idx, value=value)

        # Save the workbook
        workbook.save(excel_file)

        # combined_df = pd.DataFrame()

        # Assign DataFrame B starting from column C in the combined DataFrame
        # combined_df = pd.concat([new_df, df], axis=1)
        # print('NEW DFFFFFF', new_df)
        # print("old DFFFF", df)
        #
        # # Export the combined DataFrame to an Excel file
        # combined_df.to_excel(r'output.xlsx', index=False)


